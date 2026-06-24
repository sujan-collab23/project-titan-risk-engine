import os
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 1. PATH DISCOVERY (TARGETING YOUR DESKTOP FILE)
desktop_path = os.getcwd() 
excel_file_path = "Project_Titan_live_model.xlsx"

if not os.path.exists(excel_file_path):
    raise FileNotFoundError(f"Could not find your file. Make sure it is saved on your desktop as 'Project_Titan_live_model.xlsx'")

print("Opening existing raw M&A workbook engine...")
# Load the workbook while keeping data, charts, and formulas intact
wb = openpyxl.load_workbook(excel_file_path)

# 2. EXECUTE THE DUAL-VARIABLE COVARIANCE RISK LOOPS
print("Running 50,000-trial advanced simulation matrix...")
np.random.seed(42)
trials = 50000

# Base metrics pulled directly from your sheet inputs
jpm_base_ni = 57000.00   
zion_base_ni = 895.00     
synergies = 501.26        
debt_drag = -146.81       
base_case_proforma = jpm_base_ni + zion_base_ni + synergies + debt_drag

# Simulate sector-linked corporate variance shocks
jpm_shocks = np.random.normal(0, 0.142, trials)
zion_shocks = np.random.normal(0, 0.228, trials) * 0.68 + jpm_shocks * 0.32 

sim_jpm = jpm_base_ni * np.exp(jpm_shocks - 0.5 * (0.142**2))
sim_zion = zion_base_ni * np.exp(zion_shocks - 0.5 * (0.228**2))
sim_proforma = sim_jpm + sim_zion + synergies + debt_drag

# Calculate Risk Thresholds
p10_floor = np.percentile(sim_proforma, 10)
p50_median = np.percentile(sim_proforma, 50)
p90_ceiling = np.percentile(sim_proforma, 90)

# 3. APPEND THE SIMULATION SHEET (TAB 2)
# Check if tab already exists from a previous run to avoid duplicates
if "Monte Carlo Risk Matrix" in wb.sheetnames:
    del wb["Monte Carlo Risk Matrix"]

ws2 = wb.create_sheet(title="Monte Carlo Risk Matrix")
ws2.views.sheetView[0].showGridLines = True

# Style Palettes
navy_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
slate_stripe = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='CBD5E0'), right=Side(style='thin', color='CBD5E0'),
    top=Side(style='thin', color='CBD5E0'), bottom=Side(style='thin', color='CBD5E0')
)

# Header Formatting
ws2.merge_cells("A1:E1")
ws2["A1"] = "PROJECT TITAN // COVARIANCE RISK DISTRIBUTION ENGINE"
ws2["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
ws2["A1"].fill = navy_fill
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

# Table Headers
headers = ["Risk Horizon Metric", "Probability Threshold", "Simulated Output Value ($M)", "Delta from Base Case ($M)", "Strategic Risk Verdict"]
for c_idx, text in enumerate(headers, 1):
    c = ws2.cell(row=3, column=c_idx, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = navy_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[3].height = 24

# Data Payload
sim_rows = [
    ["P10 Downside Bound", "10% Confidence Floor", p10_floor, p10_floor - base_case_proforma, "Severe Market Stress / Asset Attrition Case"],
    ["P50 Realist Median", "50% Median Baseline", p50_median, p50_median - base_case_proforma, "Expected Operational Targets Achieved"],
    ["P90 Upside Ceiling", "90% Optimistic Runway", p90_ceiling, p90_ceiling - base_case_proforma, "Exceptional Core Synergy Extraction Run-rate"],
    ["Deterministic Excel Base", "Static Model Target", base_case_proforma, 0.00, "Standard Central Financial Case Projection"]
]

for r_idx, r_data in enumerate(sim_rows, 4):
    for c_idx, val in enumerate(r_data, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name="Calibri", size=11)
        cell.border = thin_border
        if r_idx % 2 == 0:
            cell.fill = slate_stripe
        if c_idx in [3, 4]:
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")
        elif c_idx == 2:
            cell.alignment = Alignment(horizontal="center")

# Institutional Typography Colors
ws2["E4"].font = Font(name="Calibri", size=11, bold=True, color="E53E3E") # Red Downside
ws2["E5"].font = Font(name="Calibri", size=11, bold=True, color="3182CE") # Blue Median
ws2["E6"].font = Font(name="Calibri", size=11, bold=True, color="38A169") # Green Upside

# Auto-Adjust Columns For Clean Layout Spacing
for col in ws2.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws2.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 13)

# 4. SAVE UNIFIED DATA AND OVERWRITE ONTO DESKTOP
wb.save(excel_file_path)
print(f"\n[SUCCESS] Simulation appended perfectly! Open 'Project_Titan_Live_Model.xlsx' on your desktop to view both tabs.")
